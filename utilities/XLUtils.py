import openpyxl

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def getColumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readData(file,sheetname,rowNo,colNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowNo,column=colNo).value

def writeData(file,sheetname,rowNum,columnNo,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rowNum,column=columnNo).value = data
    workbook.save(file)




